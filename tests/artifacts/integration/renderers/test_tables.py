"""Integration tests for table_renderer — no real LLM calls.

Tests verify:
  - render() produces .xlsx, .html, and N .csv files.
  - XLSX: opens with openpyxl; sheet count == table count; header styled bold + fill.
  - HTML: contains <table class="sortable"> for each table; section titles present.
  - CSV: correct row counts; headers present; one file per table.
  - render() returns all file paths.
  - Edge cases: single table, empty rows, numeric values.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict

import pytest

pytestmark = pytest.mark.integration


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def single_table_schema() -> Dict[str, Any]:
    return {
        "tables": [
            {
                "title": "Model Comparison",
                "columns": ["Model", "Provider", "Context (k)", "Cost/M tokens"],
                "rows": [
                    ["GPT-4o", "OpenAI", 128, "$5.00"],
                    ["Claude 3 Sonnet", "Anthropic", 200, "$3.00"],
                    ["Gemini 1.5 Pro", "Google", 1000, "$3.50"],
                ],
                "caption": "Prices as of 2025-Q1",
            }
        ]
    }


@pytest.fixture
def multi_table_schema() -> Dict[str, Any]:
    return {
        "tables": [
            {
                "title": "Countries",
                "columns": ["Country", "Capital", "Population (M)"],
                "rows": [
                    ["France", "Paris", 68],
                    ["Germany", "Berlin", 84],
                    ["Italy", "Rome", 59],
                ],
                "caption": "Source: World Bank 2024",
            },
            {
                "title": "Programming Languages",
                "columns": ["Language", "Year", "Paradigm"],
                "rows": [
                    ["Python", 1991, "multi"],
                    ["Rust", 2015, "systems"],
                    ["TypeScript", 2012, "OOP/FP"],
                    ["Go", 2009, "concurrent"],
                ],
                "caption": None,
            },
            {
                "title": "Empty Table",
                "columns": ["Col A", "Col B"],
                "rows": [],
                "caption": None,
            },
        ]
    }


# ---------------------------------------------------------------------------
# Unified render()
# ---------------------------------------------------------------------------

def test_render_returns_list(tmp_path: Path, single_table_schema: dict) -> None:
    from open_notebook.artifacts.renderers.table_renderer import render

    paths = render(single_table_schema, tmp_path, "tables")
    assert isinstance(paths, list)
    assert len(paths) >= 3  # xlsx + html + 1 csv


def test_render_all_files_exist(tmp_path: Path, multi_table_schema: dict) -> None:
    from open_notebook.artifacts.renderers.table_renderer import render

    paths = render(multi_table_schema, tmp_path, "report")
    for p in paths:
        assert p.exists(), f"Expected file missing: {p}"


def test_render_extensions(tmp_path: Path, multi_table_schema: dict) -> None:
    from open_notebook.artifacts.renderers.table_renderer import render

    paths = render(multi_table_schema, tmp_path, "data")
    exts = {p.suffix for p in paths}
    assert ".xlsx" in exts
    assert ".html" in exts
    assert ".csv" in exts


# ---------------------------------------------------------------------------
# XLSX tests
# ---------------------------------------------------------------------------

def test_xlsx_opens_with_openpyxl(tmp_path: Path, single_table_schema: dict) -> None:
    import openpyxl

    from open_notebook.artifacts.renderers.table_renderer import render_xlsx

    out = render_xlsx(single_table_schema, tmp_path, "tables")
    assert out.exists()
    wb = openpyxl.load_workbook(str(out))
    assert len(wb.sheetnames) >= 1


def test_xlsx_sheet_count_matches_tables(tmp_path: Path, multi_table_schema: dict) -> None:
    import openpyxl

    from open_notebook.artifacts.renderers.table_renderer import render_xlsx

    out = render_xlsx(multi_table_schema, tmp_path, "multi")
    wb = openpyxl.load_workbook(str(out))
    expected = len(multi_table_schema["tables"])
    assert len(wb.sheetnames) == expected, (
        f"Expected {expected} sheets, got {len(wb.sheetnames)}: {wb.sheetnames}"
    )


def test_xlsx_header_row_present(tmp_path: Path, single_table_schema: dict) -> None:
    import openpyxl

    from open_notebook.artifacts.renderers.table_renderer import render_xlsx

    out = render_xlsx(single_table_schema, tmp_path, "s")
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    header_values = [ws.cell(1, c).value for c in range(1, 5)]
    assert "Model" in header_values
    assert "Provider" in header_values


def test_xlsx_header_is_bold(tmp_path: Path, single_table_schema: dict) -> None:
    import openpyxl

    from open_notebook.artifacts.renderers.table_renderer import render_xlsx

    out = render_xlsx(single_table_schema, tmp_path, "s")
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    first_header_cell = ws.cell(1, 1)
    assert first_header_cell.font.bold, "Header cell should be bold"


def test_xlsx_header_has_fill(tmp_path: Path, single_table_schema: dict) -> None:
    import openpyxl

    from open_notebook.artifacts.renderers.table_renderer import render_xlsx

    out = render_xlsx(single_table_schema, tmp_path, "s")
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    cell = ws.cell(1, 1)
    # PatternFill should be set (not empty / None fill type)
    fill = cell.fill
    assert fill is not None
    assert fill.fill_type not in (None, "none"), (
        f"Header cell fill_type should not be None/none. Got: {fill.fill_type}"
    )


def test_xlsx_data_rows(tmp_path: Path, single_table_schema: dict) -> None:
    import openpyxl

    from open_notebook.artifacts.renderers.table_renderer import render_xlsx

    out = render_xlsx(single_table_schema, tmp_path, "s")
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    # Row 2 should contain first data row
    assert ws.cell(2, 1).value == "GPT-4o"
    assert ws.cell(2, 2).value == "OpenAI"


def test_xlsx_numeric_values_preserved(tmp_path: Path, multi_table_schema: dict) -> None:
    import openpyxl

    from open_notebook.artifacts.renderers.table_renderer import render_xlsx

    out = render_xlsx(multi_table_schema, tmp_path, "m")
    wb = openpyxl.load_workbook(str(out))
    # Second sheet is Programming Languages
    ws = wb.worksheets[1]
    year_val = ws.cell(2, 2).value  # row 2 = first data row, col 2 = Year
    # Should be stored as a number or string representation
    assert str(year_val) in ("1991", "1991.0"), f"Year not preserved: {year_val!r}"


def test_xlsx_sheet_names_from_titles(tmp_path: Path, multi_table_schema: dict) -> None:
    import openpyxl

    from open_notebook.artifacts.renderers.table_renderer import render_xlsx

    out = render_xlsx(multi_table_schema, tmp_path, "m")
    wb = openpyxl.load_workbook(str(out))
    names_lower = [n.lower() for n in wb.sheetnames]
    assert any("countries" in n for n in names_lower)
    assert any("programming" in n for n in names_lower)


# ---------------------------------------------------------------------------
# HTML tests
# ---------------------------------------------------------------------------

def test_html_is_valid_file(tmp_path: Path, single_table_schema: dict) -> None:
    from open_notebook.artifacts.renderers.table_renderer import render_html

    out = render_html(single_table_schema, tmp_path, "tables")
    assert out.exists()
    assert out.suffix == ".html"
    content = out.read_text(encoding="utf-8")
    assert "<!DOCTYPE html>" in content


def test_html_has_table_element(tmp_path: Path, single_table_schema: dict) -> None:
    from open_notebook.artifacts.renderers.table_renderer import render_html

    out = render_html(single_table_schema, tmp_path, "tables")
    content = out.read_text(encoding="utf-8")
    assert "<table" in content
    assert "</table>" in content


def test_html_sortable_class(tmp_path: Path, multi_table_schema: dict) -> None:
    from open_notebook.artifacts.renderers.table_renderer import render_html

    out = render_html(multi_table_schema, tmp_path, "report")
    content = out.read_text(encoding="utf-8")
    assert 'class="sortable"' in content


def test_html_table_count(tmp_path: Path, multi_table_schema: dict) -> None:
    from open_notebook.artifacts.renderers.table_renderer import render_html

    out = render_html(multi_table_schema, tmp_path, "report")
    content = out.read_text(encoding="utf-8")
    table_count = content.count("<table")
    expected = len(multi_table_schema["tables"])
    assert table_count == expected, (
        f"Expected {expected} <table> elements, got {table_count}"
    )


def test_html_section_titles(tmp_path: Path, multi_table_schema: dict) -> None:
    from open_notebook.artifacts.renderers.table_renderer import render_html

    out = render_html(multi_table_schema, tmp_path, "report")
    content = out.read_text(encoding="utf-8")
    assert "Countries" in content
    assert "Programming Languages" in content


def test_html_header_cells(tmp_path: Path, single_table_schema: dict) -> None:
    from open_notebook.artifacts.renderers.table_renderer import render_html

    out = render_html(single_table_schema, tmp_path, "s")
    content = out.read_text(encoding="utf-8")
    assert "<th>" in content
    assert "Model" in content
    assert "Provider" in content


def test_html_caption_present(tmp_path: Path, single_table_schema: dict) -> None:
    from open_notebook.artifacts.renderers.table_renderer import render_html

    out = render_html(single_table_schema, tmp_path, "s")
    content = out.read_text(encoding="utf-8")
    assert "Prices as of 2025-Q1" in content


def test_html_zebra_css(tmp_path: Path, single_table_schema: dict) -> None:
    from open_notebook.artifacts.renderers.table_renderer import render_html

    out = render_html(single_table_schema, tmp_path, "s")
    content = out.read_text(encoding="utf-8")
    assert "nth-child" in content  # zebra-stripe CSS selector


def test_html_sort_script(tmp_path: Path, single_table_schema: dict) -> None:
    from open_notebook.artifacts.renderers.table_renderer import render_html

    out = render_html(single_table_schema, tmp_path, "s")
    content = out.read_text(encoding="utf-8")
    assert "<script>" in content
    assert "sortTable" in content or "sort" in content.lower()


def test_html_xss_escaping(tmp_path: Path) -> None:
    """User-controlled data should be HTML-escaped."""
    from open_notebook.artifacts.renderers.table_renderer import render_html

    schema = {
        "tables": [
            {
                "title": "<script>alert(1)</script>",
                "columns": ["A<B", "C&D"],
                "rows": [["<b>test</b>", "x&y"]],
                "caption": None,
            }
        ]
    }
    out = render_html(schema, tmp_path, "xss_test")
    content = out.read_text(encoding="utf-8")
    # Script tag should be escaped
    assert "<script>alert" not in content
    assert "&lt;script&gt;" in content or "&amp;" in content or "&lt;" in content


# ---------------------------------------------------------------------------
# CSV tests
# ---------------------------------------------------------------------------

def test_csv_count_matches_tables(tmp_path: Path, multi_table_schema: dict) -> None:
    from open_notebook.artifacts.renderers.table_renderer import render_csv

    csv_paths = render_csv(multi_table_schema, tmp_path, "report")
    expected = len(multi_table_schema["tables"])
    assert len(csv_paths) == expected, (
        f"Expected {expected} CSV files, got {len(csv_paths)}"
    )


def test_csv_files_exist(tmp_path: Path, multi_table_schema: dict) -> None:
    from open_notebook.artifacts.renderers.table_renderer import render_csv

    paths = render_csv(multi_table_schema, tmp_path, "report")
    for p in paths:
        assert p.exists(), f"CSV file missing: {p}"


def test_csv_headers(tmp_path: Path, single_table_schema: dict) -> None:
    import csv as csv_mod

    from open_notebook.artifacts.renderers.table_renderer import render_csv

    paths = render_csv(single_table_schema, tmp_path, "s")
    assert paths
    with paths[0].open(newline="", encoding="utf-8") as f:
        reader = csv_mod.reader(f)
        headers = next(reader)
    assert "Model" in headers
    assert "Provider" in headers


def test_csv_row_count(tmp_path: Path, single_table_schema: dict) -> None:
    import csv as csv_mod

    from open_notebook.artifacts.renderers.table_renderer import render_csv

    paths = render_csv(single_table_schema, tmp_path, "s")
    with paths[0].open(newline="", encoding="utf-8") as f:
        rows = list(csv_mod.reader(f))
    # 1 header + 3 data rows
    assert len(rows) == 4, f"Expected 4 rows (1 header + 3 data), got {len(rows)}"


def test_csv_utf8_encoding(tmp_path: Path) -> None:
    from open_notebook.artifacts.renderers.table_renderer import render_csv

    schema = {
        "tables": [
            {
                "title": "Unicode",
                "columns": ["Name", "Value"],
                "rows": [["Ångström", "1.0 Å"], ["Café", "☕"]],
                "caption": None,
            }
        ]
    }
    paths = render_csv(schema, tmp_path, "unicode")
    content = paths[0].read_text(encoding="utf-8")
    assert "Ångström" in content
    assert "Café" in content


def test_csv_stems_contain_index(tmp_path: Path, multi_table_schema: dict) -> None:
    from open_notebook.artifacts.renderers.table_renderer import render_csv

    paths = render_csv(multi_table_schema, tmp_path, "report")
    filenames = [p.name for p in paths]
    # Files should be numbered 1_, 2_, 3_
    assert any("_1_" in n or "_1_" in n for n in filenames)
    assert any("_2_" in n for n in filenames)
