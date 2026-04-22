"""Table renderers for DataTablesSchema.

Outputs XLSX, HTML, and CSV (one file per table) from a DataTablesSchema.

Public API
----------
render(schema, output_dir: Path, stem: str) -> List[Path]
    Returns all produced file paths.

Individual renderers (also exported):
  render_xlsx(schema, output_dir, stem) -> Path
  render_html(schema, output_dir, stem) -> Path
  render_csv(schema, output_dir, stem)  -> List[Path]
"""
from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any, Dict, List


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _to_dict(schema) -> Dict[str, Any]:
    if isinstance(schema, dict):
        return schema
    return schema.model_dump()


def _safe_stem(text: str) -> str:
    """Slugify a string for use as a filename stem."""
    slug = re.sub(r"[^A-Za-z0-9_\-]", "_", text.strip())
    return slug[:40] or "table"


# ---------------------------------------------------------------------------
# XLSX renderer
# ---------------------------------------------------------------------------

def render_xlsx(schema, output_dir: Path, stem: str) -> Path:
    """Write all tables to a single .xlsx file (one sheet per table).

    Sheet names are derived from table titles (truncated to 31 chars — Excel
    limit).  The header row is styled bold with a blue fill.  Column widths
    are auto-fitted to the maximum content length.

    Returns the path to the written .xlsx file.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    data = _to_dict(schema)
    tables = data.get("tables", [])

    wb = openpyxl.Workbook()
    # Remove default sheet
    if wb.active:
        wb.remove(wb.active)

    HEADER_FILL = PatternFill("solid", fgColor="2E74B5")
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    BODY_FONT = Font(name="Calibri", size=11)

    for tbl in tables:
        if isinstance(tbl, dict):
            title = str(tbl.get("title", "Table"))
            columns = list(tbl.get("columns") or [])
            rows = list(tbl.get("rows") or [])
            caption = tbl.get("caption")
        else:
            title = str(tbl.title)
            columns = list(tbl.columns)
            rows = list(tbl.rows)
            caption = getattr(tbl, "caption", None)

        # Excel sheet names max 31 chars; strip illegal chars
        sheet_name = re.sub(r"[\\/*?\[\]:]", "", title)[:31] or "Table"
        ws = wb.create_sheet(title=sheet_name)

        # Header row
        for ci, col in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=ci, value=str(col))
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Data rows
        for ri, row in enumerate(rows, start=2):
            padded = list(row) + [""] * max(0, len(columns) - len(row))
            padded = padded[:len(columns)]
            for ci, val in enumerate(padded, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = BODY_FONT
                cell.alignment = Alignment(vertical="top", wrap_text=False)

        # Caption below data
        if caption:
            caption_row = len(rows) + 3
            ws.cell(row=caption_row, column=1, value=str(caption)).font = Font(
                name="Calibri", size=10, italic=True
            )

        # Auto-fit column widths
        for ci, col in enumerate(columns, start=1):
            col_letter = get_column_letter(ci)
            max_len = len(str(col))
            for ri_data in range(len(rows)):
                row_data = rows[ri_data]
                if ci - 1 < len(row_data):
                    max_len = max(max_len, len(str(row_data[ci - 1])))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

        # Freeze header row
        ws.freeze_panes = "A2"

    output_dir.mkdir(parents=True, exist_ok=True)
    out = output_dir / f"{stem}.xlsx"
    wb.save(str(out))
    return out


# ---------------------------------------------------------------------------
# HTML renderer
# ---------------------------------------------------------------------------

_HTML_HEAD = """\
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{title}</title>
<style>
  body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
         font-size: 14px; color: #111827; padding: 24px; max-width: 1200px;
         margin: 0 auto; }}
  h2 {{ color: #1f4a7d; border-bottom: 2px solid #2e74b5; padding-bottom: 6px; }}
  p.caption {{ font-size: 12px; color: #6b7280; font-style: italic; margin-top: 4px; }}
  table {{ border-collapse: collapse; width: 100%; margin-bottom: 8px; }}
  thead tr {{ background: #2e74b5; color: #fff; }}
  thead th {{ padding: 8px 12px; text-align: left; cursor: pointer;
              user-select: none; white-space: nowrap; }}
  thead th:hover {{ background: #1f4a7d; }}
  tbody tr:nth-child(even) {{ background: #f0f7ff; }}
  tbody tr:hover {{ background: #dbeafe; }}
  td {{ padding: 6px 12px; border-bottom: 1px solid #e5e7eb; }}
  .sort-asc::after  {{ content: " ▲"; font-size: 10px; }}
  .sort-desc::after {{ content: " ▼"; font-size: 10px; }}
</style>
</head>
<body>
"""

_SORT_SCRIPT = """\
<script>
(function() {
  function sortTable(th) {
    var table = th.closest('table');
    var tbody = table.querySelector('tbody');
    var colIdx = Array.from(th.parentNode.children).indexOf(th);
    var asc = th.classList.toggle('sort-asc');
    th.parentNode.querySelectorAll('th').forEach(function(t) {
      if (t !== th) { t.classList.remove('sort-asc', 'sort-desc'); }
    });
    if (!asc) { th.classList.add('sort-desc'); }
    var rows = Array.from(tbody.querySelectorAll('tr'));
    rows.sort(function(a, b) {
      var av = a.cells[colIdx] ? a.cells[colIdx].textContent.trim() : '';
      var bv = b.cells[colIdx] ? b.cells[colIdx].textContent.trim() : '';
      var an = parseFloat(av), bn = parseFloat(bv);
      if (!isNaN(an) && !isNaN(bn)) return asc ? an - bn : bn - an;
      return asc ? av.localeCompare(bv) : bv.localeCompare(av);
    });
    rows.forEach(function(r) { tbody.appendChild(r); });
  }
  document.querySelectorAll('table.sortable thead th').forEach(function(th) {
    th.addEventListener('click', function() { sortTable(th); });
  });
})();
</script>
"""


def _html_escape(text: str) -> str:
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def render_html(schema, output_dir: Path, stem: str) -> Path:
    """Write all tables to a single sortable HTML file.

    Returns the path to the written .html file.
    """
    data = _to_dict(schema)
    tables = data.get("tables", [])

    doc_title = stem.replace("_", " ").title()
    parts = [_HTML_HEAD.format(title=_html_escape(doc_title))]
    parts.append(f"<h1>{_html_escape(doc_title)}</h1>\n")

    for tbl in tables:
        if isinstance(tbl, dict):
            title = str(tbl.get("title", "Table"))
            columns = list(tbl.get("columns") or [])
            rows = list(tbl.get("rows") or [])
            caption = tbl.get("caption")
        else:
            title = str(tbl.title)
            columns = list(tbl.columns)
            rows = list(tbl.rows)
            caption = getattr(tbl, "caption", None)

        parts.append(f"<section>\n<h2>{_html_escape(title)}</h2>\n")
        parts.append('<table class="sortable" data-sortable>\n<thead>\n<tr>\n')
        for col in columns:
            parts.append(f"  <th>{_html_escape(str(col))}</th>\n")
        parts.append("</tr>\n</thead>\n<tbody>\n")

        for row in rows:
            padded = list(row) + [""] * max(0, len(columns) - len(row))
            padded = padded[:len(columns)]
            parts.append("<tr>\n")
            for val in padded:
                parts.append(f"  <td>{_html_escape(str(val))}</td>\n")
            parts.append("</tr>\n")

        parts.append("</tbody>\n</table>\n")
        if caption:
            parts.append(f'<p class="caption">{_html_escape(str(caption))}</p>\n')
        parts.append("</section>\n\n")

    parts.append(_SORT_SCRIPT)
    parts.append("</body>\n</html>\n")

    output_dir.mkdir(parents=True, exist_ok=True)
    out = output_dir / f"{stem}.html"
    out.write_text("".join(parts), encoding="utf-8")
    return out


# ---------------------------------------------------------------------------
# CSV renderer (one file per table)
# ---------------------------------------------------------------------------

def render_csv(schema, output_dir: Path, stem: str) -> List[Path]:
    """Write one CSV file per table.

    File names: <stem>_<N>_<slug>.csv

    Returns a list of written file paths.
    """
    data = _to_dict(schema)
    tables = data.get("tables", [])

    output_dir.mkdir(parents=True, exist_ok=True)
    paths: List[Path] = []

    for i, tbl in enumerate(tables):
        if isinstance(tbl, dict):
            title = str(tbl.get("title", f"table_{i}"))
            columns = list(tbl.get("columns") or [])
            rows = list(tbl.get("rows") or [])
        else:
            title = str(tbl.title)
            columns = list(tbl.columns)
            rows = list(tbl.rows)

        filename = f"{stem}_{i + 1}_{_safe_stem(title)}.csv"
        out = output_dir / filename

        with out.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerow(columns)
            for row in rows:
                padded = list(row) + [""] * max(0, len(columns) - len(row))
                writer.writerow([str(v) for v in padded[:len(columns)]])

        paths.append(out)

    return paths


# ---------------------------------------------------------------------------
# Unified render()
# ---------------------------------------------------------------------------

def render(schema, output_dir: Path, stem: str) -> List[Path]:
    """Render a DataTablesSchema to XLSX + HTML + CSV files.

    Parameters
    ----------
    schema:
        DataTablesSchema Pydantic model or compatible dict.
    output_dir:
        Directory where output files are written.  Created if needed.
    stem:
        Base filename (without extension) for XLSX and HTML outputs.

    Returns
    -------
    List[Path]
        All produced file paths: [xlsx, html, *csv_files].
    """
    paths: List[Path] = []
    paths.append(render_xlsx(schema, output_dir, stem))
    paths.append(render_html(schema, output_dir, stem))
    paths.extend(render_csv(schema, output_dir, stem))
    return paths
